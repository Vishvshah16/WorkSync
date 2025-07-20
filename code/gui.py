from tkinter import Tk, Label, Entry, Text, Button, Toplevel, messagebox, Frame
from tkinter import ttk
from db_handler import fetch_tasks_for_today, add_task_to_db, update_task_in_db, save_email_to_db, fetch_email_configured
from datetime import date, datetime, timezone
import tzlocal
import smtplib
from email.message import EmailMessage
import base64
from openpyxl import Workbook, load_workbook
import os


today = date.today()

# Format: DD-MM-YYYY
current_date = today.strftime("%d %B, %Y, %A")

def export_today_tasks_to_excel(file_path="daily_tasks_details.xlsx"):
    tasks = fetch_tasks_for_today()
    if not tasks:
        print("No tasks found for today.")
        return
    headers = ["ID", "Task Detail", "Current Status", "Help Taken", "Completion Time", "Reason", "Deployment Status", "Comments"]
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        # Remove the default sheet if starting fresh
        default_sheet = wb.active
        wb.remove(default_sheet)

    sheet_name = current_date
    if os.path.exists(file_path):
        wb = load_workbook(file_path)

        if sheet_name in wb.sheetnames:
            std = wb[sheet_name]
            wb.remove(std)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)

    for task in tasks:
        ws.append(task)

    wb.save(file_path)

def send_email():
    tasks = fetch_tasks_for_today()

    if not tasks:
        body = "No tasks found for today."
    else:
        lines = ["Today's Task Summary:\n"]
        for i, task in enumerate(tasks, 1):
            lines.append(f"Task {i}:")
            lines.append(f"  Task Detail: {task[1]}")
            lines.append(f"  Current Status: {task[2]}")
            lines.append(f"  Help Taken: {task[3]}")
            lines.append(f"  Time Taken: {task[4]}")
            lines.append(f"  Reason If Not Completed: {task[5]}")
            lines.append(f"  Deployment Status: {task[6]}")
            lines.append(f"  Additonal Comments: {task[7]}\n")
        body = "\n".join(lines)
    regards = "\n\nThanks"
    body = body+regards
    
    print(body)

    email_details = fetch_email_configured()
    print(email_details)
    for i, email in enumerate(email_details, 1):
        sender = email[0]
        password = email[1]
        receiver = email[2]
        cc  = email[3]
        subject = email[4]

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = receiver
    msg["CC"] = cc
    msg.set_content(body)

    password = base64.b64decode(password)
    decoded_password = password.decode('utf-8')

    # Send via SMTP (Example: Gmail)
    try:
        smtpserver = smtplib.SMTP('smtp.office365.com', 587)
        smtpserver.ehlo()
        smtpserver.starttls()
        smtpserver.ehlo()
        smtpserver.login(sender, decoded_password)
        msg = msg.as_string()
        smtpserver.sendmail(sender, receiver, msg)
        smtpserver.close()
    except Exception as e:
        print("Email send failed:", e)
        return False

    messagebox.showinfo("Success", f"Task details sent successfully!")

def open_edit_popup(task):
    """Open a popup window to edit the selected task."""

    task_id = task["task_id"] 
    task_detail = task["task_detail"] 
    current_status = task["current_status"] 
    help_taken = task["help_taken"] 
    completion_time = task["completion_time"] 
    reason = task["reason"] 
    deployment_status = task["deployment_status"] 
    comments = task["comments"] 
    

    popup = Toplevel()
    popup.title(f"Edit Task")
    popup.configure(bg="#faedcd")
    width, height = 450, 350
    popup.geometry(f"{width}x{height}") 

    popup.update_idletasks()  # Ensure geometry info is up-to-date
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()
    popup.iconbitmap("./assets/icon.ico")
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    popup.geometry(f"{width}x{height}+{x}+{y}")

    # Prevent resizing
    popup.resizable(False, False)

    # Title label at the top
    Label(popup, text="Update Task Details", bg="#faedcd", foreground="#432818").grid(row=0, column=0, columnspan=2, pady=(15, 10))


    Label(popup, text="Task Detail:", background="#faedcd", foreground="#432818").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    task_detail_entry = Entry(popup, width=40)
    task_detail_entry.insert(0, task_detail)
    task_detail_entry.grid(row=1, column=1, padx=10, pady=5)

    Label(popup, text="Current Status:", background="#faedcd", foreground="#432818").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    current_status_entry = Entry(popup, width=40)
    current_status_entry.insert(0, current_status)
    current_status_entry.grid(row=2, column=1, padx=10, pady=5)

    Label(popup, text="Help Taken:", background="#faedcd", foreground="#432818").grid(row=3, column=0, padx=10, pady=5, sticky="e")
    help_taken_entry = Entry(popup, width=40)
    help_taken_entry.insert(0, help_taken)
    help_taken_entry.grid(row=3, column=1, padx=10, pady=5)

    Label(popup, text="Completion Time:", background="#faedcd", foreground="#432818").grid(row=4, column=0, padx=10, pady=5, sticky="e")
    completion_time_entry = Entry(popup, width=40)
    completion_time_entry.insert(0, completion_time or "")
    completion_time_entry.grid(row=4, column=1, padx=10, pady=5)

    Label(popup, text="Reason if not completed:", background="#faedcd", foreground="#432818").grid(row=5, column=0, padx=10, pady=5, sticky="e")
    reason_entry = Entry(popup, width=40)
    reason_entry.insert(0, reason)
    reason_entry.grid(row=5, column=1, padx=10, pady=5)

    Label(popup, text="Deployment Status:", background="#faedcd", foreground="#432818").grid(row=6, column=0, padx=10, pady=5, sticky="e")
    deployment_status_entry = Entry(popup, width=40)
    deployment_status_entry.insert(0, deployment_status)
    deployment_status_entry.grid(row=6, column=1, padx=10, pady=5)

    Label(popup, text="Comments:", background="#faedcd", foreground="#432818").grid(row=7, column=0, padx=10, pady=5, sticky="e")
    comments_text = Entry(popup, width=40)
    comments_text.insert(0, comments)
    comments_text.grid(row=7, column=1, padx=10, pady=5)

    Button(popup, text="Update Task", background="#d4a373", command=lambda: submit_task_update(
        task_id, task_detail_entry, current_status_entry, help_taken_entry,
        completion_time_entry, reason_entry, deployment_status_entry, comments_text, popup)
    ).grid(row=8, column=0, columnspan=2, pady=10)


def submit_task_update(task_id, task_detail_entry, current_status_entry, help_taken_entry,
                       completion_time_entry, reason_entry, deployment_status_entry, comments_text, popup):
    """Submit the updated task to the database."""
    task_detail = task_detail_entry.get()
    current_status = current_status_entry.get()
    help_taken = help_taken_entry.get()
    completion_time = completion_time_entry.get()
    reason = reason_entry.get()
    deployment_status = deployment_status_entry.get()
    comments = comments_text.get()

    update_task_in_db(task_id, task_detail, current_status, help_taken, completion_time, reason, deployment_status, comments)
    messagebox.showinfo("Success", f"Task details updated successfully!")
    popup.destroy()
    refresh_task_table()


def refresh_task_table():
    for row in task_table.get_children():
        task_table.delete(row)

    tasks = fetch_tasks_for_today()
    for task in tasks:
        task_id, task_detail, current_status, help_taken, completion_time, reason, deployment_status, comments = task
        task_table.insert("", "end", values=(task_id, task_detail, current_status, help_taken, completion_time, reason, deployment_status, comments))

def edit_selected_task():
    selected_item = task_table.selection()
    if not selected_item:
        messagebox.showwarning("No Selection", "Please select a task to edit.")
        return

    # Get task data from the selected row
    selected_values = task_table.item(selected_item[0], "values")

    task_data = {
        "task_id": selected_values[0],
        "task_detail": selected_values[1],
        "current_status": selected_values[2],
        "help_taken": selected_values[3],
        "completion_time": selected_values[4],
        "reason": selected_values[5],
        "deployment_status": selected_values[6],
        "comments": selected_values[7],
    }
    open_edit_popup(task_data)


def display_task_table_and_form(root):
    global task_table

    # Task Table
    Label(root, text="Today's Tasks", background="#faedcd", foreground="#000000").grid(row=0, column=0, columnspan=2, pady=10)
    Label(root, text=f"Date: {current_date}", anchor="e", background="#faedcd", foreground="#000000").grid(row=0, column=1, sticky="e", padx=20, pady=10)

    task_table = ttk.Treeview(root, columns=("Sr. No.","Detail", "Status", "Help", "Time", "Reason", "Deployment", "Comments"), show="headings", height=8)
    task_table.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

    task_table.heading("Sr. No.", text="Sr. No.")
    task_table.column("Sr. No.", width=150, anchor="w")

    task_table.heading("Detail", text="Detail")
    task_table.column("Detail", width=150, anchor="w")

    task_table.heading("Status", text="Status")
    task_table.column("Status", width=100, anchor="center")

    task_table.heading("Help", text="Help")
    task_table.column("Help", width=100, anchor="center")

    task_table.heading("Time", text="Time")
    task_table.column("Time", width=80, anchor="center")

    task_table.heading("Reason", text="Reason")
    task_table.column("Reason", width=120, anchor="w")

    task_table.heading("Deployment", text="Deployment")
    task_table.column("Deployment", width=120, anchor="w")
    
    task_table.heading("Comments", text="Comments")
    task_table.column("Comments", width=120, anchor="w")

    # Populate table with today's tasks
    refresh_task_table()

    Button(root, text="Edit Task", background="#d4a373", foreground="#000000",command=edit_selected_task).grid(row=2, column=0, columnspan=2, pady=10)
    
    # Add Task Form
    Label(root, text="Add New Task", background="#faedcd", foreground="#000000").grid(row=3, column=0, columnspan=2, pady=10)

    form_frame = Frame(root)
    form_frame.configure(bg="#faedcd")
    form_frame.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

    Label(form_frame, text="Task Detail:", background="#faedcd", foreground="#000000").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    task_detail_entry = Entry(form_frame, width=40)
    task_detail_entry.grid(row=0, column=1, padx=5, pady=5)

    Label(form_frame, text="Current Status:", background="#faedcd", foreground="#000000").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    current_status_entry = Entry(form_frame, width=40)
    current_status_entry.grid(row=1, column=1, padx=5, pady=5)

    Label(form_frame, text="Help Taken:", background="#faedcd", foreground="#000000").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    help_taken_entry = Entry(form_frame, width=40)
    help_taken_entry.grid(row=2, column=1, padx=5, pady=5)

    Label(form_frame, text="Completion Time:", background="#faedcd", foreground="#000000").grid(row=3, column=0, sticky="e", padx=5, pady=5)
    completion_time_entry = Entry(form_frame, width=40)
    completion_time_entry.grid(row=3, column=1, padx=5, pady=5)

    Label(form_frame, text="Reason if not completed:", background="#faedcd", foreground="#000000").grid(row=4, column=0, sticky="e", padx=5, pady=5)
    reason_entry = Entry(form_frame, width=40)
    reason_entry.grid(row=4, column=1, padx=5, pady=5)

    Label(form_frame, text="Deployment Status:", background="#faedcd", foreground="#000000").grid(row=5, column=0, sticky="e", padx=5, pady=5)
    deployment_status_entry = Entry(form_frame, width=40)
    deployment_status_entry.grid(row=5, column=1, padx=5, pady=5)

    Label(form_frame, text="Additional Comments:", background="#faedcd", foreground="#000000").grid(row=6, column=0, sticky="e", padx=5, pady=5)
    comments_text = Entry(form_frame, width=40)
    comments_text.grid(row=6, column=1, padx=5, pady=5)

    Button(root, text="Add Task", background="#d4a373", command=lambda: add_new_task(
        task_detail_entry, current_status_entry, help_taken_entry, completion_time_entry, reason_entry, deployment_status_entry, comments_text)
    ).grid(row=5, column=0, columnspan=2, pady=10)
    
    Button(root, text="Get Data", background="#d4a373", command=lambda: export_today_tasks_to_excel()).grid(row=6, column=0, columnspan=2, pady=10)


def add_new_task(task_detail_entry, current_status_entry, help_taken_entry, completion_time_entry, reason_entry, deployment_status_entry, comments_text):
    """Add a new task to the database."""
    task_detail = task_detail_entry.get()
    current_status = current_status_entry.get()
    help_taken = help_taken_entry.get()
    completion_time = completion_time_entry.get()
    reason = reason_entry.get()
    deployment_status = deployment_status_entry.get()
    comments = comments_text.get()
    if not task_detail:
        messagebox.showwarning("Input Error", "Task detail is required.")
        return
    
    current_utc_date = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    utc_dt = datetime.strptime(current_utc_date, "%Y-%m-%dT%H:%M:%S.%fZ")
    utc_dt = utc_dt.replace(tzinfo=timezone.utc)
    local_tz = tzlocal.get_localzone()
    created_at = utc_dt.astimezone(local_tz).strftime("%Y-%m-%d ")
    
    add_task_to_db(task_detail, current_status, help_taken, completion_time, reason, deployment_status, comments,created_at)
    messagebox.showinfo("Success", "New Task added successfully!")

    # Clear the form and refresh the table
    task_detail_entry.delete(0, "end")
    current_status_entry.delete(0, "end")
    help_taken_entry.delete(0, "end")
    completion_time_entry.delete(0, "end")
    reason_entry.delete(0, "end")
    deployment_status_entry.delete(0, "end")
    comments_text.delete(0, "end")
    refresh_task_table()


def create_gui():
    root = Tk()
    root.title("WorkSync Task Tracker")
    root.resizable(False, False)
    
    root.iconbitmap("./assets/icon.ico")
    window_width = 965
    window_height = 680

    root.update_idletasks()

    # Get screen size
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calculate position
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    # Set the geometry with the calculated position
    root.geometry(f"{window_width}x{window_height}+{x}+{y}") # Optional: Adjust the window size
    root.configure(bg="#faedcd")
    display_task_table_and_form(root)

    root.mainloop()

def create_gui_email(root_callback):
    root = Tk()
    root.title("Email Configuration")
    root.resizable(False, False)
    
    root.iconbitmap("./assets/icon.ico")
    window_width = 430
    window_height = 300
    
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    root.geometry(f"{window_width}x{window_height}+{x}+{y}") # Optional: Adjust the window size
    root.configure(bg="#faedcd")

    Label(root, text="Email Details", background="#faedcd", foreground="#000000").grid(row=0, column=0, columnspan=2, pady=10)

    form_frame = Frame(root)
    form_frame.configure(bg="#faedcd")
    form_frame.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

    Label(form_frame, text="Sender Email:", background="#faedcd", foreground="#000000").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    sender_email_entry = Entry(form_frame, width=40)
    sender_email_entry.grid(row=0, column=1, padx=5, pady=5)

    Label(form_frame, text="Sender App Password:", background="#faedcd", foreground="#000000").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    app_password_entry = Entry(form_frame, width=40)
    app_password_entry.grid(row=1, column=1, padx=5, pady=5)

    Label(form_frame, text="Receiver Email:", background="#faedcd", foreground="#000000").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    receiver_email_entry = Entry(form_frame, width=40)
    receiver_email_entry.grid(row=2, column=1, padx=5, pady=5)

    Label(form_frame, text="CC:", background="#faedcd", foreground="#000000").grid(row=3, column=0, sticky="e", padx=5, pady=5)
    cc_email_entry = Entry(form_frame, width=40)
    cc_email_entry.grid(row=3, column=1, padx=5, pady=5)

    Label(form_frame, text="Subject:", background="#faedcd", foreground="#000000").grid(row=4, column=0, sticky="e", padx=5, pady=5)
    subject_entry = Entry(form_frame, width=40)
    subject_entry.grid(row=4, column=1, padx=5, pady=5)

    def custom_error(title, message, bg_color="#b08968", btn_color="#7f5539"):
        error_win = Toplevel()
        error_win.title(title)
        error_win.configure(bg=bg_color)
        window_width = 300
        window_height = 120
        error_win.resizable(False, False)
        error_win.iconbitmap("./assets/icon.ico")
        screen_width = error_win.winfo_screenwidth()
        screen_height = error_win.winfo_screenheight()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)

        error_win.geometry(f"{window_width}x{window_height}+{x}+{y}")
        # Message label
        msg_label = Label(error_win, text=message, bg=bg_color, fg="black")
        msg_label.pack(pady=20)

        # OK button
        ok_button = Button(
            error_win,
            text="OK",
            command=error_win.destroy,
            bg=btn_color,
            fg="white",
            activebackground="#63372c",
            relief="flat",
            width=10
        )
        ok_button.pack(pady=5)

        # Make sure the window is modal (blocks interaction with other windows)
        error_win.grab_set()
        error_win.transient()
        
    def on_submit():
        sender = sender_email_entry.get()
        password = app_password_entry.get()
        receiver = receiver_email_entry.get()
        cc = cc_email_entry.get()
        subject = subject_entry.get()
        
        if not sender or not receiver or not subject or not password:
            custom_error("Error", "Required details are missing for email.")
            return

        save_email_to_db(sender,password, receiver, cc, subject)
        root.destroy()
        root_callback()

    submit_button = Button(form_frame, text="Submit", background="#d4a373", foreground="#000000",command=on_submit)
    submit_button.grid(row=5, column=0, columnspan=2, pady=10)

    root.mainloop()
